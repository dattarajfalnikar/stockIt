


from flask import Flask, render_template, request
# from flask import
from flask_migrate import Migrate
import xlrd
import ipdb
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
import json
# import requests
# from urllib import  request
# from bs4 import BeautifulSoup
import yfinance as yf
import datetime
from io import BytesIO
from flask_mail import Mail, Message
import pandas_datareader.data as web
from nsetools import Nse
import yahoo_fin.stock_info as y_fin
import math

app = Flask(__name__)

app.config['MAIL_SERVER']='smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_DEFAULT_SENDER'] = 'stockitdev@gmail.com'
app.config['MAIL_USERNAME'] = 'stockitdev@gmail.com'
app.config['MAIL_PASSWORD'] = 'stockitdev@pucsd'
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True

mail = Mail(app)

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:123@localhost/stockit'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)

NSE = '.NS'
BSE = '.BO'


class Stocks(db.Model):

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    company_name = db.Column(db.String(200))
    symbol = db.Column(db.String(100))
    country = db.Column(db.String(100), default='India')
    yf_symbol = db.Column(db.String(100), default=None)
    face_value = db.Column(db.String(10), default='0')
    exchange = db.Column(db.String(50), default='NSE')
    stock_details = db.Column(db.JSON, default={})

class DailyStockDataYahooFin(db.Model):

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    company_name = db.Column(db.String(200))
    symbol = db.Column(db.String(100))
    country = db.Column(db.String(100), default='India')
    prev_close = db.Column(db.String(100))
    open = db.Column(db.String(100))
    high = db.Column(db.String(100))
    low = db.Column(db.String(100))
    close = db.Column(db.String(100))
    volume = db.Column(db.String(100))
    date = db.Column(db.Date, default=datetime.date.today())


class DailyStockData(db.Model):

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    company_name = db.Column(db.String(200))
    symbol = db.Column(db.String(100))
    country = db.Column(db.String(100), default='India')
    prev_close_price = db.Column(db.Float, default=None)
    open_price = db.Column(db.Float, default=None)
    day_high_price = db.Column(db.Float, default=None)
    day_low_price = db.Column(db.Float, default=None)
    close_price = db.Column(db.Float, default=None)
    last_price = db.Column(db.Float, default=None)
    average_price = db.Column(db.Float, default=None)
    upper_band = db.Column(db.Float, default=None) #upper circuit
    lower_band = db.Column(db.Float, default=None) #lower circuit
    high52 = db.Column(db.Float, default=None)
    low52 = db.Column(db.Float, default=None)
    purpose = db.Column(db.String(300), default=None)
    record_date = db.Column(db.String(100), default=None)
    ex_date = db.Column(db.String(100), default=None)
    face_value = db.Column(db.Float, default=None)
    delivery_quantity = db.Column(db.Float, default=None)
    quantity_traded = db.Column(db.Float, default=None)
    delivery_to_traded_quantity_percentage = db.Column(db.Float, default=None)
    total_traded_volume = db.Column(db.Float, default=None)
    total_buy_quantity = db.Column(db.Float, default=None)
    total_sell_quantity = db.Column(db.Float, default=None)
    daily_percentage_change = db.Column(db.Float, default=None)
    one_day_percentage_change = db.Column(db.Float, default=None)
    css_status = db.Column(db.String(100), default=None)
    date = db.Column(db.Date, default=datetime.date.today())


class Dailylogs(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    company_name = db.Column(db.String(200))
    symbol = db.Column(db.String(100))
    error = db.Column(db.String(1000))
    date = db.Column(db.Date, default=datetime.date.today())


class DailyStockAnalysis(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    company_name = db.Column(db.String(200))
    symbol = db.Column(db.String(100))
    prev_close = db.Column(db.String(100))
    open = db.Column(db.String(100))
    high = db.Column(db.String(100))
    low = db.Column(db.String(100))
    close = db.Column(db.String(100))
    volume = db.Column(db.String(100))
    date = db.Column(db.Date, default=datetime.date.today())
    today_percentage_change = db.Column(db.Float, default=0.00)
    one_day_percentage_change = db.Column(db.Float, default=0.00)


class UserEmails(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    email = db.Column(db.String(100))


def send_daily_analysis_email(trix, send_to=[], name='StockIt Analysis-', user_msg=None, html=None):
    if not user_msg:
        msg = Message('Stockit Daily Analysis ' + str(datetime.date.today()))
    else:
        msg = Message(user_msg + str(datetime.date.today()))

    if not html:
        msg.html = '<div> Hello Investor,' \
                   '<br><br> Please find attached xl-sheet about Todays Analysis. </div>' \
                   ' <br><div> Note: Data Is based on NSE Equity, You might see minor Differences in stock price </div>'
    else:
        msg.html = html

    msg.recipients = send_to
    msg.attach(''.join((name, str(datetime.date.today()), '.xlsx')),
               'application/MICROSOFT_EXCEL', trix.getvalue())
    mail.send(msg)


@app.route('/')
def home_page():
    email_list = ['dattarajfalnikar@gmail.com', '007.sarju@gmail.com', 'deven0000007@gmail.com',
                  'ishancshinde23@gmail.com',
                  'chavare29@gmail.com',
                  'vaibhav.mukadam@gmail.com',
                  ]
    # user_list = []
    #
    # for individual_email in email_list:
    #     user_list.append(UserEmails(email=individual_email))
    #
    # db.session.add_all(user_list)
    # db.session.commit()

    return render_template('home.html', name='StockIT')


def get_stock_details():
    """
    The function is just filter stock_details and provide filtered data, for long term investment,
    predictions form Yahoo Finance about stock target prices.

    :return: stock data list
    """
    data = []
    for stock in Stocks.query.all():
        data_dict = {}
        print(stock.symbol, '::', stock.company_name)
        if stock.stock_details:
            stock_data = json.loads(stock.stock_details)
            print(stock_data)
            try:
                if math.isnan(float(stock_data.get('1y Target Est'))) or not stock_data.get('1y Target Est'):
                    data_dict['one_year_target_est'] = 0.0
                else:
                    data_dict['one_year_target_est'] = stock_data.get('1y Target Est')
                data_dict['52_week_range'] = stock_data.get('52 Week Range')
                data_dict['avg_volume'] = stock_data.get('Avg. Volume')
                data_dict['beta_5y_monthly'] = stock_data.get('Beta (5Y Monthly)')
                data_dict["day's_range"] = stock_data.get("Day's Range")
                data_dict['ex_dividend_date'] = stock_data.get('Ex-Dividend Date')
                data_dict['forward_dividend_yield'] = stock_data.get('Forward Dividend & Yield')
                data_dict['Market Cap'] = stock_data.get('Market Cap')
                data_dict['pe_ratio'] = stock_data.get('PE Ratio (TTM)')
                data_dict['eps_ttm'] = stock_data.get('EPS (TTM)')
                data_dict['earnings_date'] = stock_data.get('Earnings Date')
                data_dict['symbol'] = stock.symbol
                data_dict['company_name'] = stock.company_name
                data_dict['one_year_target_est_percentage'] = 0.0

                data_dict['prev_close'] = 0.0
                if stock_data.get('Previous Close') and not math.isnan(float(stock_data.get('Previous Close'))):
                    data_dict['prev_close'] = stock_data.get('Previous Close')

                if data_dict.get('one_year_target_est') and data_dict.get('prev_close'):
                    data_dict['one_year_target_est_percentage'] = ((data_dict.get('one_year_target_est') - (
                        data_dict.get('prev_close'))) * 100) / data_dict.get('prev_close')
                print(data_dict)
            except ValueError:
                print('Decoding Json Failed!!')

            if data_dict:
                data.append(data_dict)
    return data


@app.route('/yahoo_stock_target')
def yahoo_stock_target():
    index = 0
    stock_details = get_stock_details()
    # ipdb.set_trace()
    stocks = sorted(stock_details, key=lambda k: k['one_year_target_est_percentage'], reverse=True)
    return render_template('one_year_target_yahoo.html', data=stocks)


@app.route('/update_stock_details')
def update_stock_details():
    index = 0
    for stock in Stocks.query.all():
        print(index, stock.symbol, '::', stock.company_name)

        try:
            data = y_fin.get_quote_table(stock.symbol + NSE)
            print(data)
            if type(data) is dict:
                stock.stock_details = json.dumps(data)
                db.session.commit()
        except ValueError:
            print('Decoding Json Failed!!')

        index += 1
    return 'Stock details Updated!'


@app.route('/dailyreportemailyahoo', methods=['GET'])
def daily_report_email_yahoo():
    # ipdb.set_trace()
    wb = Workbook()
    sheet = wb.active

    sheet.append(('Company Name', 'Symbol', 'Prev_Close', 'Open', 'High', 'Low', 'Close',
                  'Daily Change in % Open/Close', 'One Day Change% prevClose/Close', 'Date'))
    for data in DailyStockAnalysis.query.filter(
            DailyStockAnalysis.date==datetime.date.today()).order_by(
        DailyStockAnalysis.today_percentage_change.desc()).all():
        sheet.append((data.company_name, data.symbol, data.prev_close, data.open, data.high, data.low, data.close,
                      data.today_percentage_change, data.one_day_percentage_change, data.date))

    xl_sheet = BytesIO()
    wb.save(xl_sheet)

    for data in UserEmails.query.all():
        send_daily_analysis_email(xl_sheet, [data.email],)
    # send_daily_analysis_email(xl_sheet, ['vaibhav.mukadam@gmail.com'],)

    return render_template('home.html', name='StockIT')


@app.route('/dailyreportemail', methods=['GET'])
def daily_report_email():
    # ipdb.set_trace()
    wb = Workbook()
    sheet = wb.active

    sheet.append(('Company Name', 'Symbol', 'Prev_Close', 'Open', 'Close', 'Upper Circuit',
                  'Daily Change in % Open/Close', 'One Day Change% prevClose/Close', 'High', 'Low', 'Lower Circuit',
                  '52 Weeks Low', '52 Weeks High', 'TradedQuantity'
                  'DeliveryQuantity', 'Ex Dividend Date', 'Purpose of Ex Date'))
    for data in DailyStockData.query.filter(
            DailyStockData.date==datetime.date.today()).order_by(
        DailyStockData.daily_percentage_change.desc()).all():
        sheet.append((data.company_name, data.symbol, data.prev_close_price, data.open_price,
                      data.close_price, data.upper_band, data.daily_percentage_change, data.one_day_percentage_change,
                      data.day_high_price, data.day_low_price, data.lower_band,
                      data.low52, data.high52, data.quantity_traded,
                      data.delivery_quantity, data.ex_date, data.purpose))

    xl_sheet = BytesIO()
    wb.save(xl_sheet)

    for data in UserEmails.query.all():
        send_daily_analysis_email(xl_sheet, [data.email],)
    # send_daily_analysis_email(xl_sheet, ['dattarajfalnikar@gmail.com'],)

    return render_template('home.html', name='StockIT')


@app.route('/dailyreportyahoo', methods=['GET'])
def daily_report_yahoo():
    # ipdb.set_trace()
    context_data = {}
    data_list = []
    for data in DailyStockAnalysis.query.filter(
            DailyStockAnalysis.date==datetime.date.today()).order_by(
        DailyStockAnalysis.today_percentage_change.desc()).all():
        display_data={}
        display_data['company_name'] = data.company_name
        display_data['symbol'] = data.symbol
        display_data['prev_close'] = data.prev_close
        display_data['open'] = data.open
        display_data['high'] = data.high
        display_data['low'] = data.low
        display_data['close'] = data.close
        display_data['today_change'] = data.today_percentage_change
        display_data['one_day_change'] = data.one_day_percentage_change
        data_list.append(display_data)
    context_data['data'] = data_list
    context_data['date'] = str(datetime.date.today())
    # ipdb.set_trace()
    return render_template('daily_report.html', name='StockIT', data=context_data)


@app.route('/dailyreport', methods=['GET', 'POST'])
def daily_report_nsetool():
    # ipdb.set_trace()
    context_data = {}
    data_list = []
    #-datetime.timedelta(days=1)
    if request.form.get('selected_date'):
        get_record_for = datetime.datetime.strptime(request.form.get('selected_date'), '%Y-%m-%d').date()
    else:
        get_record_for = datetime.date.today()

    db_query = None
    if request.form.get('filter') == 'sort_by_daily_change':
        db_query = DailyStockData.query.filter(
            DailyStockData.date == get_record_for).order_by(
            DailyStockData.daily_percentage_change.desc()).all()
    elif request.form.get('filter') == 'this_month_ex_date':
        month = get_record_for.strftime('%b')
        if get_record_for > datetime.date.today():
            get_record_for = datetime.date.today() - datetime.timedelta(days=1)

        db_query = DailyStockData.query.filter(
            DailyStockData.date == get_record_for,
            DailyStockData.ex_date.contains('-'.join((month, get_record_for.strftime('%y'))))
        ).order_by(DailyStockData.close_price).all()
    else:
        db_query = DailyStockData.query.filter(
            DailyStockData.date==get_record_for).order_by(
        DailyStockData.delivery_to_traded_quantity_percentage.desc()).all()

    for data in db_query:
        # DailyStockData.daily_percentage_change.desc()).all():
        display_data = {}
        display_data['company_name'] = data.company_name
        display_data['symbol'] = data.symbol
        display_data['prev_close'] = data.prev_close_price
        display_data['open'] = data.open_price
        display_data['high'] = data.day_high_price
        display_data['low'] = data.day_low_price
        display_data['close'] = data.close_price
        display_data['day_change'] = data.daily_percentage_change
        display_data['one_day_change'] = data.one_day_percentage_change
        display_data['high52'] = data.high52
        display_data['low52'] = data.low52
        display_data['quantity_traded'] = data.quantity_traded
        display_data['delivery_quantity'] = data.delivery_quantity
        display_data['total_delivery_percentage'] = data.delivery_to_traded_quantity_percentage
        display_data['purpose'] = data.purpose
        display_data['upper_circuit'] = data.upper_band
        display_data['lower_circuit'] = data.lower_band
        display_data['ex_date'] = data.ex_date
        data_list.append(display_data)
    if not data_list:
        context_data = {'err': 'No Data Found!',}
    context_data['data'] = data_list
    context_data['date'] = get_record_for
    return render_template('daily_report.html', name='StockIT', data=context_data)


@app.route('/dailyanalysisyahoo')
def daily_analysis():

    all_data = DailyStockData.query.filter(DailyStockData.date == datetime.date.today()).all()
    # all_data = DailyStockData.query.all()
    daily_analysis_data = []
    for data in all_data:
      try:
          open_close_diff = ((float(data.close) - float(data.open)) * 100) / float(data.open)

          if data.prev_close:
              prevclose_close_diff = ((float(data.close) - float(data.prev_close)) * 100) / float(data.prev_close)
          else:
              prevclose_close_diff = None

          daily_analysis_data.append(DailyStockAnalysis(symbol=data.symbol, company_name=data.company_name,
                                                        prev_close=data.prev_close,
                                                        open=data.open, close=data.close,
                                                        high=data.high, low=data.low,
                                                        today_percentage_change=open_close_diff,
                                                        one_day_percentage_change=prevclose_close_diff))

      except:
          print('Error')
          return 'Internal Server Error, 500'

    db.session.add_all(daily_analysis_data)
    db.session.commit()
    return 'Analysis Complteted'


@app.route('/dailydatayahoo', methods=['GET'])
def daily_data_yahoo():
    if datetime.date.weekday(datetime.date.today()) not in [5, 6]:
        stocks = Stocks.query.all()
        all_data = []
        all_logs = []
        if datetime.date.weekday(datetime.date.today()) == 0:
            days = 3
        else:
            days = 1
        start_date = datetime.date.today() - datetime.timedelta(days=days)
        for stock in stocks:
            try:
                data = yf.download(stock.symbol + NSE, actions=True,
                                   start=start_date)
                # data = web.DataReader(stock.symbol + NSE, 'yahoo',  start_date )
                print(stock.symbol, data)
                if len(data) == 2:
                    prev_close = data.Close.values[0]
                    close_val = data.Close.values[1]
                    open_val = data.Open.values[1]
                    high_val = data.High.values[1]
                    low_val = data.Low.values[1]
                    volume = data.Low.values[1]
                else:
                    pre_close =  DailyStockData.query.filter(
                        DailyStockData.symbol.like(stock.symbol), DailyStockData.date.like(datetime.date.today()-datetime.timedelta(days=1))).first()
                    if pre_close:
                        prev_close = pre_close.prev_close
                    else:
                        prev_close = None
                    close_val = data.Close.values[0]
                    open_val = data.Open.values[0]
                    high_val = data.High.values[0]
                    low_val = data.Low.values[0]
                    volume = data.Low.values[0]
                    # date_ = data.Date.values[1]

                all_data.append(DailyStockData(company_name=stock.company_name, symbol=stock.symbol,
                                               prev_close=prev_close,
                                               open=open_val, close=close_val,
                                               high=high_val, low=low_val, volume=volume))

            except IndexError as e:
                all_logs.append(Dailylogs(symbol=stock.symbol, company_name=stock.company_name, error=e))
                print(stock.symbol, '======================================', stock.company_name)

        db.session.add_all(all_data)
        db.session.add_all(all_logs)
        db.session.commit()

        return 'Data Added Successfully'
    return 'Today is Weekend'


@app.route('/dailydata', methods=['GET'])
def daily_data_nsetool():
    # datetime.datetime.strptime(data.get('secDate').split(' ')[0], '%d-%B-%Y').date()   #datetime get from nse tool
    nse = Nse()
    index = 1
    # ipdb.set_trace()
    if datetime.date.weekday(datetime.date.today()) not in [5, 6]:
        # stocks = Stocks.query.all()
        all_data = []
        all_logs = []
        secdate = None
        daily_p_change = None
        #        # for stock in Stocks.query.filter(Stocks.symbol.like('BKMINDST')):
        for stock in Stocks.query.all():
            stock_data = None
            print(stock.symbol, '======================================', stock.company_name)
            try:
                stock_data = nse.get_quote(stock.symbol)
            except Exception as e:
                print(e)
            try:
                if stock_data:
                    print(index, stock_data)
                    openprice = stock_data.get('open')
                    closeprice = stock_data.get('closePrice')
                    if stock_data.get('secDate'):
                        secdate = datetime.datetime.strptime(stock_data.get('secDate').split(' ')[0], '%d-%b-%Y').date()
                    if openprice and closeprice:
                        daily_p_change = ((closeprice - openprice) * 100) / openprice

                    all_data.append(DailyStockData(
                                                   company_name=stock.company_name,
                                                   symbol=stock.symbol,
                                                   prev_close_price=stock_data.get('previousClose'),
                                                   open_price=openprice,
                                                   close_price=closeprice,
                                                   day_low_price=stock_data.get('dayLow'),
                                                   day_high_price=stock_data.get('dayHigh'),
                                                   last_price=stock_data.get('lastPrice'),
                                                   average_price=stock_data.get('averagePrice'),
                                                   upper_band=stock_data.get('pricebandupper'),
                                                   lower_band=stock_data.get('pricebandlower'),
                                                   high52=stock_data.get('high52'),
                                                   low52=stock_data.get('low52'),
                                                   purpose=stock_data.get('purpose'),
                                                   record_date=stock_data.get('recordDate'),
                                                   ex_date=stock_data.get('exDate'),
                                                   delivery_quantity=stock_data.get('deliveryQuantity'),
                                                   delivery_to_traded_quantity_percentage=stock_data.get(
                                                       'deliveryToTradedQuantity'),
                                                   face_value=stock_data.get('faceValue'),
                                                   quantity_traded=stock_data.get('quantityTraded'),
                                                   total_traded_volume=stock_data.get('totalTradedVolume'),
                                                   total_buy_quantity=stock_data.get('totalBuyQuantity'),
                                                   total_sell_quantity=stock_data.get('totalSellQuantity'),
                                                   daily_percentage_change=daily_p_change,
                                                   one_day_percentage_change=stock_data.get('pChange'),
                                                   css_status=stock_data.get('css_status_desc'),
                                                   date=secdate,
                                                   ))
                index += 1
            except IndexError as e:
                all_logs.append(Dailylogs(symbol=stock.symbol, company_name=stock.company_name, error=e))
                print(stock.symbol, '+++++++++++++++++++++++++++', stock.company_name)

        if not all_data:
            return 'No Data is Added'

        db.session.add_all(all_data)
        db.session.add_all(all_logs)
        db.session.commit()

        return 'Data Added Successfully'
    return 'Today is Weekend'


@app.route('/uploadfile/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        excel_sheet = request.files['file']
        wb = load_workbook(excel_sheet)
        ws = wb[wb.sheetnames[0]]
        test = []
        final_list = []
        for index, row in enumerate(ws.values):
            if index > 0 and row[1] is not None:
                data = Stocks(company_name=row[1], symbol=row[0], face_value=row[7])
                test.append([row[0], row[1]])
                final_list.append(data)
        db.session.add_all(final_list)
        db.session.commit()

        return json.dumps(test)
    return render_template('upload_file.html', name='StockIT Upload')


if __name__  ==  "__main__":
    db.create_all()
    app.run(debug=True)
